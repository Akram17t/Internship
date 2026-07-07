from __future__ import annotations

import base64
import binascii
import hashlib
import hmac
import json
import re
import secrets
import sys
import threading
import time
import uuid
from datetime import datetime
from datetime import timedelta
from datetime import timezone
from io import BytesIO
from pathlib import Path
from urllib.parse import quote, unquote

from fastapi import FastAPI, Header, HTTPException
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from backend.settings import get_env, load_capstone_env

ROOT_DIR = Path(__file__).resolve().parents[2]
load_capstone_env()
CREW_SRC_DIR = ROOT_DIR / "backend" / "researcher_crew" / "src"
if str(CREW_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(CREW_SRC_DIR))

app = FastAPI(title="ICS Knowledge Assistant API", version="1.0.0")
FRONTEND_DIR = ROOT_DIR / "frontend" / "web"
ASSETS_DIR = FRONTEND_DIR / "assets"
EMBEDDABLE_EXTENSIONS = {".pdf", ".docx", ".txt"}
LIBRARY_EXTENSIONS = EMBEDDABLE_EXTENSIONS | {".xlsx"}
MAX_DOCUMENT_BYTES = 25 * 1024 * 1024
ADMIN_SESSION_TTL = timedelta(hours=12)
MAX_CONVERSATIONS = 50
MAX_CONVERSATION_TURNS = 5
MAX_CONVERSATION_MESSAGES = MAX_CONVERSATION_TURNS * 2
MAX_CONVERSATION_CONTEXT_CHARS = 3200
CONVERSATION_TTL = timedelta(days=1)
CONVERSATION_LOCK = threading.Lock()
FAQ_LOCK = threading.Lock()
REINDEX_LOCK = threading.Lock()
ADMIN_CONFIG_LOCK = threading.Lock()

def _new_admin_config_template() -> dict[str, str]:
    # Buat template config admin yang sengaja kosong.
    return {
        "email": "",
        "password": "",
        "name": "Admin",
        "session_secret": secrets.token_hex(32),
    }

class QueryRequest(BaseModel):
    question: str = Field(..., min_length=3, description="Natural language question from the user.")
    conversation_id: str | None = Field(default=None, description="Client conversation identifier.")


class CitationResponse(BaseModel):
    id: int
    source: str
    page: int | None = None
    section: str | None = None
    chunk_id: int | None = None
    download_url: str | None = None


class FormDownloadResponse(BaseModel):
    name: str
    display_name: str
    download_url: str


class QueryResponse(BaseModel):
    answer: str
    citations: list[CitationResponse] = Field(default_factory=list)
    form_downloads: list[FormDownloadResponse] = Field(default_factory=list)
    conversation_id: str


class FAQItem(BaseModel):
    id: str
    question: str
    answer: str
    source: str = ""
    source_url: str = ""
    suggested_query: str
    citations: list[CitationResponse] = Field(default_factory=list)
    image_url: str = ""
    updated_at: str | None = None


class AdminFAQPayload(BaseModel):
    question: str = Field(..., min_length=3)


class AdminFAQResponse(BaseModel):
    message: str
    item: FAQItem | None = None


class AdminLoginPayload(BaseModel):
    email: str = Field(..., min_length=3)
    password: str = Field(..., min_length=1)


class AdminLoginResponse(BaseModel):
    email: str
    name: str
    token: str
    expires_at: str


class LibraryItem(BaseModel):
    name: str
    relative_path: str
    display_name: str
    doc_type: str
    document_kind: str
    is_embeddable: bool
    size_bytes: int
    updated_at: str
    download_url: str


class AdminDocumentPayload(BaseModel):
    filename: str = Field(..., min_length=1)
    content_base64: str = Field(..., min_length=1)
    replace_path: str | None = None


class AdminDocumentResponse(BaseModel):
    message: str
    requires_reindex: bool = False
    item: LibraryItem | None = None


class AdminReindexResponse(BaseModel):
    message: str


def _get_data_dir() -> Path:
    # Tentukan folder data backend dari konfigurasi env.
    raw_dir = get_env("DATA_DIR", "backend/data")
    path = Path(raw_dir)
    if not path.is_absolute():
        path = ROOT_DIR / path
    return path


def _document_kind_for_path(path: Path) -> str:
    # Klasifikasikan file tersimpan sebagai form, SOP, atau dokumen umum.
    name = path.stem.lower()
    if path.suffix.lower() == ".xlsx" or name.startswith("form"):
        return "form"
    if name.startswith("sop"):
        return "sop"
    return "document"


def _is_embeddable_path(path: Path) -> bool:
    # Kembalikan True jika file perlu masuk ke vector DB.
    return path.suffix.lower() in EMBEDDABLE_EXTENSIONS


def _to_library_item(path: Path, data_dir: Path) -> LibraryItem:
    # Ubah file tersimpan menjadi bentuk respons library admin.
    relative_path = path.relative_to(data_dir).as_posix()
    stat = path.stat()
    display_name = path.stem.replace("_", " ").replace("-", " ").strip() or path.name
    return LibraryItem(
        name=path.name,
        relative_path=relative_path,
        display_name=display_name.title(),
        doc_type=path.suffix.lower().lstrip("."),
        document_kind=_document_kind_for_path(path),
        is_embeddable=_is_embeddable_path(path),
        size_bytes=stat.st_size,
        updated_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
        download_url=f"/api/documents/{quote(relative_path, safe='/')}",
    )


def _iter_library_items() -> list[LibraryItem]:
    # Daftar semua file yang didukung di folder data saat ini.
    data_dir = _get_data_dir()
    if not data_dir.exists():
        return []

    items: list[LibraryItem] = []
    for path in sorted(data_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in LIBRARY_EXTENSIONS:
            continue
        if path.name.startswith("~$"):  # skip Excel lock files
            continue

        items.append(_to_library_item(path, data_dir))

    return items


def _format_form_display_name(path: Path) -> str:
    # Ubah nama file form mentah menjadi label yang lebih rapi.
    return (
        path.stem.replace("Form - ", "")
        .replace("(Template)", "")
        .replace("_", " ")
        .strip()
    )


def _form_download_response(path: Path, data_dir: Path) -> FormDownloadResponse:
    # Bentuk payload download publik untuk file form.
    relative_path = path.relative_to(data_dir).as_posix()
    return FormDownloadResponse(
        name=path.name,
        display_name=_format_form_display_name(path),
        download_url=f"/api/documents/{quote(relative_path, safe='/')}",
    )


def _iter_form_downloads() -> list[FormDownloadResponse]:
    # Daftar semua template form yang bisa diunduh.
    data_dir = _get_data_dir()
    if not data_dir.exists():
        return []

    forms: list[FormDownloadResponse] = []
    for path in _iter_form_paths(data_dir):
        forms.append(_form_download_response(path, data_dir))
    return forms


def _iter_form_paths(data_dir: Path | None = None) -> list[Path]:
    # Kumpulkan semua path form xlsx sambil lewati lock Excel sementara.
    data_dir = data_dir or _get_data_dir()
    if not data_dir.exists():
        return []

    paths: list[Path] = []
    for path in sorted(data_dir.rglob("*.xlsx")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        paths.append(path)
    return paths


def _available_form_catalog(forms: list[FormDownloadResponse]) -> str:
    # Ubah daftar form menjadi katalog yang bisa dipilih AI.
    if not forms:
        return "[]"

    return json.dumps(
        [
            {
                "name": form.name,
                "display_name": form.display_name,
            }
            for form in forms
        ],
        ensure_ascii=False,
    )


def _form_lookup_keys(form: FormDownloadResponse) -> set[str]:
    # Buat key pencocokan longgar untuk satu pilihan form.
    return {
        form.name.strip().lower(),
        form.display_name.strip().lower(),
        Path(form.name).stem.strip().lower(),
        Path(form.display_name).stem.strip().lower(),
    }


def _selected_form_downloads(
    selected_names: list[str],
    forms: list[FormDownloadResponse],
) -> list[FormDownloadResponse]:
    # Cocokkan nama form pilihan AI ke payload download yang nyata.
    if not selected_names or not forms:
        return []

    lookup: dict[str, FormDownloadResponse] = {}
    for form in forms:
        for key in _form_lookup_keys(form):
            if key:
                lookup[key] = form

    selected: list[FormDownloadResponse] = []
    seen_names: set[str] = set()
    for raw_name in selected_names:
        form = lookup.get(raw_name.strip().lower())
        if form is None or form.name in seen_names:
            continue
        selected.append(form)
        seen_names.add(form.name)
    return selected


def _answer_has_supported_form_context(answer: str) -> bool:
    # Sembunyikan form download jika jawabannya sebenarnya fallback tanpa sumber.
    normalized = " ".join(answer.lower().split())
    unsupported_markers = (
        "tidak tersedia dalam dokumen",
        "tidak tersedia di dokumen",
        "tidak disebutkan",
        "tidak dinyatakan",
        "tidak ada ketentuan",
        "tidak dapat menemukan informasi terkait hal tersebut",
        "tidak ada informasi",
        "tidak memuat",
        "tidak mencakup",
        "tidak menjelaskan",
        "tanpa menyebutkan",
        "tidak ditemukan",
        "belum tersedia",
        "belum dapat dikonfirmasi",
        "tidak dapat dikonfirmasi",
        "dokumen yang terindeks tidak",
        "dokumen terindeks tidak",
    )
    return not any(marker in normalized for marker in unsupported_markers)


def _admin_email() -> str:
    # Ambil email admin yang terkonfigurasi.
    return _load_admin_config()["email"].strip().lower()


def _admin_name() -> str:
    # Ambil nama tampilan admin yang terkonfigurasi.
    return _load_admin_config()["name"].strip() or "Admin"


def _admin_password() -> str:
    # Ambil password admin yang terkonfigurasi.
    return _load_admin_config()["password"]


def _admin_session_secret() -> str:
    # Ambil secret penanda tangan token sesi admin.
    return _load_admin_config()["session_secret"]


def _base64url_encode(value: bytes) -> str:
    # Encode bytes ke base64 URL-safe tanpa padding.
    return base64.urlsafe_b64encode(value).rstrip(b"=").decode("ascii")


def _base64url_decode(value: str) -> bytes:
    # Decode base64 URL-safe yang mungkin tanpa padding.
    padding = "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode((value + padding).encode("ascii"))


def _sign_admin_payload(payload: str) -> str:
    # Buat signature HMAC untuk payload sesi admin.
    return hmac.new(
        _admin_session_secret().encode("utf-8"),
        payload.encode("ascii"),
        hashlib.sha256,
    ).hexdigest()


def _create_admin_token(email: str) -> tuple[str, datetime]:
    # Buat token sesi admin bertanda tangan dengan waktu kedaluwarsa.
    expires_at = datetime.now(timezone.utc) + ADMIN_SESSION_TTL
    payload = _base64url_encode(
        json.dumps(
            {"email": email, "exp": int(expires_at.timestamp())},
            separators=(",", ":"),
        ).encode("utf-8")
    )
    return f"{payload}.{_sign_admin_payload(payload)}", expires_at


def _verify_admin_token(authorization: str) -> str:
    # Validasi bearer token dan kembalikan email admin jika valid.
    scheme, _, token = authorization.strip().partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=401, detail="Admin login required.")

    payload, separator, signature = token.partition(".")
    if not separator or not payload or not signature:
        raise HTTPException(status_code=401, detail="Invalid admin session.")
    if not hmac.compare_digest(signature, _sign_admin_payload(payload)):
        raise HTTPException(status_code=401, detail="Invalid admin session.")

    try:
        data = json.loads(_base64url_decode(payload).decode("utf-8"))
    except (ValueError, json.JSONDecodeError, binascii.Error) as error:
        raise HTTPException(status_code=401, detail="Invalid admin session.") from error

    email = str(data.get("email", "")).strip().lower()
    expires_at = int(data.get("exp", 0))
    if email != _admin_email() or expires_at <= int(time.time()):
        raise HTTPException(status_code=401, detail="Admin session expired.")
    return email


def _require_admin(authorization: str) -> str:
    # Lindungi endpoint dengan verifikasi token admin.
    return _verify_admin_token(authorization)


def _resolve_document_path(document_path: str) -> Path:
    # Tentukan path dokumen relatif sambil mencegah path traversal.
    data_dir = _get_data_dir().resolve()
    resolved_path = (data_dir / unquote(document_path)).resolve()

    try:
        resolved_path.relative_to(data_dir)
    except ValueError as error:
        raise HTTPException(status_code=400, detail="Invalid document path.") from error

    return resolved_path


def _decode_document(content_base64: str) -> bytes:
    # Decode file upload base64 dan terapkan batas ukuran.
    try:
        payload = base64.b64decode(content_base64, validate=True)
    except (binascii.Error, ValueError) as error:
        raise HTTPException(status_code=400, detail="Invalid document payload.") from error

    if not payload:
        raise HTTPException(status_code=400, detail="Document cannot be empty.")
    if len(payload) > MAX_DOCUMENT_BYTES:
        raise HTTPException(status_code=413, detail="Document is too large.")
    return payload


def _get_cache_dir() -> Path:
    # Tentukan folder cache lokal untuk file state JSON.
    raw_dir = get_env("CONVERSATION_CACHE_DIR", "backend/cache")
    path = Path(raw_dir)
    if not path.is_absolute():
        path = ROOT_DIR / path
    return path


def _get_conversation_file() -> Path:
    # Kembalikan path file cache percakapan.
    return _get_cache_dir() / "conversations.json"


def _get_faq_file() -> Path:
    # Kembalikan path file cache FAQ.
    return _get_cache_dir() / "faqs.json"


def _get_admin_file() -> Path:
    # Kembalikan path file config admin.
    return _get_cache_dir() / "admin.json"


def _save_admin_config(config: dict[str, str]) -> None:
    # Simpan JSON config admin ke disk.
    cache_dir = _get_cache_dir()
    cache_dir.mkdir(parents=True, exist_ok=True)
    _get_admin_file().write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _load_admin_config() -> dict[str, str]:
    # Muat config admin dari disk dan isi default aman yang masih kurang.
    with ADMIN_CONFIG_LOCK:
        path = _get_admin_file()
        if not path.exists():
            config = _new_admin_config_template()
            _save_admin_config(config)
            return config

        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            data = {}

        if not isinstance(data, dict):
            data = {}

        session_secret = str(data.get("session_secret") or "").strip()
        if not session_secret:
            session_secret = secrets.token_hex(32)

        config = {
            "email": str(data.get("email") or "").strip().lower(),
            "password": str(data.get("password") or ""),
            "name": str(data.get("name") or "Admin").strip() or "Admin",
            "session_secret": session_secret,
        }
        if data != config:
            _save_admin_config(config)
        return config


def _clean_conversation_id(value: str | None) -> str:
    # Sanitasi conversation ID dari client atau buat yang baru.
    if not value:
        return uuid.uuid4().hex

    cleaned = "".join(char for char in value if char.isalnum() or char in {"-", "_"})
    if 8 <= len(cleaned) <= 80:
        return cleaned
    return uuid.uuid4().hex


def _parse_conversation_timestamp(value: object) -> datetime | None:
    # Parse timestamp percakapan tersimpan jika valid.
    if not isinstance(value, str) or not value.strip():
        return None

    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def _prune_expired_conversations(
    conversations: dict[str, list[dict[str, object]]],
) -> dict[str, list[dict[str, object]]]:
    # Buang thread percakapan yang pesan terbarunya sudah lewat TTL.
    cutoff = datetime.now() - CONVERSATION_TTL
    pruned: dict[str, list[dict[str, object]]] = {}

    for conversation_id, messages in conversations.items():
        if not isinstance(messages, list) or not messages:
            continue

        latest_timestamp: datetime | None = None
        for message in reversed(messages):
            if not isinstance(message, dict):
                continue
            latest_timestamp = _parse_conversation_timestamp(message.get("created_at"))
            if latest_timestamp is not None:
                break

        if latest_timestamp is None or latest_timestamp < cutoff:
            continue

        pruned[conversation_id] = messages

    return pruned


def _load_conversations() -> dict[str, list[dict[str, object]]]:
    # Muat dan pangkas riwayat percakapan dari disk.
    path = _get_conversation_file()
    if not path.exists():
        return {}

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}

    if not isinstance(data, dict):
        return {}

    conversations = {
        str(key): value
        for key, value in data.items()
        if isinstance(value, list)
    }
    return _prune_expired_conversations(conversations)


def _save_conversations(conversations: dict[str, list[dict[str, object]]]) -> None:
    # Simpan riwayat percakapan yang sudah dibatasi dan dipangkas.
    cache_dir = _get_cache_dir()
    cache_dir.mkdir(parents=True, exist_ok=True)

    active_conversations = _prune_expired_conversations(conversations)
    trimmed_items = list(active_conversations.items())[-MAX_CONVERSATIONS:]
    payload = {key: value[-MAX_CONVERSATION_MESSAGES:] for key, value in trimmed_items}
    _get_conversation_file().write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _get_conversation_context(conversation_id: str) -> str:
    # Ubah turn terbaru menjadi context teks untuk rewrite query.
    with CONVERSATION_LOCK:
        messages = _load_conversations().get(conversation_id, [])

    context_lines: list[str] = []
    for message in messages[-MAX_CONVERSATION_MESSAGES:]:
        role = "User" if message.get("role") == "user" else "Assistant"
        content = str(message.get("content", "")).strip()
        if content:
            context_lines.append(f"{role}: {content}")

    return "\n".join(context_lines)[-MAX_CONVERSATION_CONTEXT_CHARS:]


def _append_conversation_turn(conversation_id: str, question: str, answer: str) -> None:
    # Tambahkan satu pasangan turn user/assistant ke cache percakapan.
    now = datetime.now().isoformat(timespec="seconds")
    with CONVERSATION_LOCK:
        conversations = _load_conversations()
        messages = conversations.setdefault(conversation_id, [])
        messages.extend(
            [
                {"role": "user", "content": question, "created_at": now},
                {"role": "assistant", "content": answer, "created_at": now},
            ]
        )
        conversations[conversation_id] = messages[-MAX_CONVERSATION_MESSAGES:]
        _save_conversations(conversations)


def _citation_download_url(source: str) -> str:
    # Buat URL download dokumen dari nama file sumber citation.
    return f"/api/documents/{quote(source, safe='')}" if source else ""


def _normalize_citation(raw_item: object, index: int) -> CitationResponse | None:
    # Normalisasi satu dict citation mentah ke model respons API.
    if not isinstance(raw_item, dict):
        return None

    source = str(raw_item.get("source", "")).strip()
    if not source:
        return None

    return CitationResponse(
        id=int(raw_item.get("id") or index + 1),
        source=source,
        page=raw_item.get("page") if isinstance(raw_item.get("page"), int) else None,
        section=str(raw_item.get("section", "")).strip() or None,
        chunk_id=raw_item.get("chunk_id") if isinstance(raw_item.get("chunk_id"), int) else None,
        download_url=str(raw_item.get("download_url", "")).strip() or _citation_download_url(source),
    )


def _normalize_citations(item: dict[str, object]) -> list[CitationResponse]:
    # Normalisasi citation dengan fallback legacy source/source_url.
    raw_citations = item.get("citations")
    if isinstance(raw_citations, list):
        citations = [
            citation
            for citation in (
                _normalize_citation(raw_item, index)
                for index, raw_item in enumerate(raw_citations)
            )
            if citation is not None
        ]
        if citations:
            return citations

    source = str(item.get("source", "")).strip()
    source_url = str(item.get("source_url", "")).strip()
    if not source:
        return []

    return [
        CitationResponse(
            id=1,
            source=source,
            download_url=source_url or _citation_download_url(source),
        )
    ]


def _normalize_faq_item(item: dict[str, object]) -> FAQItem | None:
    # Normalisasi satu record FAQ tersimpan ke model API.
    question = str(item.get("question", "")).strip()
    answer = str(item.get("answer", "")).strip()
    if not question or not answer:
        return None

    citations = _normalize_citations(item)
    source = str(item.get("source", "")).strip()
    source_url = str(item.get("source_url", "")).strip()
    if citations and not source:
        source = citations[0].source
    if citations and not source_url:
        source_url = citations[0].download_url or ""

    return FAQItem(
        id=str(item.get("id") or uuid.uuid4().hex),
        question=question,
        answer=answer,
        source=source,
        source_url=source_url,
        suggested_query=str(item.get("suggested_query", "")).strip() or question,
        citations=citations,
        image_url=str(item.get("image_url", "")).strip(),
        updated_at=str(item.get("updated_at", "")).strip() or None,
    )


def _load_faqs() -> list[FAQItem]:
    # Muat item FAQ dari cache JSON lokal.
    path = _get_faq_file()
    if not path.exists():
        return []

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []

    if not isinstance(data, list):
        return []

    return [
        item
        for item in (_normalize_faq_item(raw_item) for raw_item in data if isinstance(raw_item, dict))
        if item is not None
    ]


def _save_faqs(items: list[FAQItem]) -> None:
    # Simpan item FAQ ke cache JSON lokal.
    cache_dir = _get_cache_dir()
    cache_dir.mkdir(parents=True, exist_ok=True)
    payload = [
        item.model_dump() if hasattr(item, "model_dump") else item.dict()
        for item in items
    ]
    _get_faq_file().write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _find_faq_index(items: list[FAQItem], faq_id: str) -> int:
    # Cari index item FAQ berdasarkan ID atau lempar 404.
    for index, item in enumerate(items):
        if item.id == faq_id:
            return index
    raise HTTPException(status_code=404, detail="FAQ not found.")


def _is_unusable_faq_answer(answer: str, citations: list[CitationResponse]) -> bool:
    # Tolak jawaban FAQ yang tidak punya evidence atau terlalu generik.
    normalized = " ".join(answer.lower().split())
    blocked_phrases = (
        "informasi ini belum tersedia",
        "informasi tersebut tidak tersedia",
        "tidak tersedia dalam dokumen",
        "tidak ditemukan dalam dokumen",
        "pertanyaan anda tidak jelas",
        "mohon berikan detail",
        "berikan detail lebih lanjut",
        "[nama perusahaan]",
    )
    return not citations or any(phrase in normalized for phrase in blocked_phrases)


def _build_faq_item(payload: AdminFAQPayload, faq_id: str | None = None) -> FAQItem:
    # Buat dan validasi satu entri FAQ dari pertanyaan.
    from researcher_crew.main import OllamaGenerationError, run_faq_crew

    question = payload.question.strip()
    try:
        answer, raw_citations = run_faq_crew(question)
    except OllamaGenerationError as error:
        raise HTTPException(status_code=502, detail=str(error)) from error
    citations = [
        CitationResponse(
            **citation,
            download_url=f"/api/documents/{quote(str(citation['source']), safe='')}",
        )
        for citation in raw_citations
    ]
    if _is_unusable_faq_answer(answer, citations):
        raise HTTPException(
            status_code=422,
            detail=(
                "FAQ tidak disimpan karena tidak ada sumber dari dokumen terindeks. "
                "Coba tulis pertanyaan yang lebih spesifik atau tambahkan dokumen yang relevan."
            ),
        )
    source = citations[0].source if citations else ""
    source_url = citations[0].download_url if citations else ""
    return FAQItem(
        id=faq_id or uuid.uuid4().hex,
        question=question,
        answer=answer,
        source=source,
        source_url=source_url or "",
        suggested_query=question,
        citations=citations,
        updated_at=datetime.now().isoformat(timespec="seconds"),
    )


if ASSETS_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(ASSETS_DIR)), name="assets")


@app.get("/health")
def health_check() -> dict[str, str]:
    # Probe sederhana untuk mengecek backend hidup.
    return {"status": "ok"}


@app.post("/api/admin/login", response_model=AdminLoginResponse)
def login_admin(payload: AdminLoginPayload) -> AdminLoginResponse:
    # Autentikasi admin lalu buat token sesi.
    if not _admin_email() or not _admin_password():
        raise HTTPException(
            status_code=503,
            detail="Admin belum dikonfigurasi. Isi email dan password di backend/cache/admin.json.",
        )

    email = payload.email.strip().lower()
    password = payload.password
    if (
        not hmac.compare_digest(email, _admin_email())
        or not hmac.compare_digest(password, _admin_password())
    ):
        raise HTTPException(status_code=401, detail="Email atau password admin salah.")

    token, expires_at = _create_admin_token(email)
    return AdminLoginResponse(
        email=email,
        name=_admin_name(),
        token=token,
        expires_at=expires_at.isoformat(timespec="seconds"),
    )


@app.post("/query", response_model=QueryResponse)
def query_knowledge_base(payload: QueryRequest) -> QueryResponse:
    # Jawab query chat dengan citation dan form pilihan AI.
    from researcher_crew.main import OllamaGenerationError, run_knowledge_crew

    conversation_id = _clean_conversation_id(payload.conversation_id)
    conversation_context = _get_conversation_context(conversation_id)
    available_forms = _iter_form_downloads()
    try:
        answer, raw_citations, selected_form_names = run_knowledge_crew(
            payload.question,
            conversation_context,
            available_forms=_available_form_catalog(available_forms),
        )
    except OllamaGenerationError as error:
        raise HTTPException(status_code=502, detail=str(error)) from error
    _append_conversation_turn(conversation_id, payload.question, answer)
    citations = [
        CitationResponse(
            **citation,
            download_url=f"/api/documents/{quote(str(citation['source']), safe='')}",
        )
        for citation in raw_citations
    ]
    form_downloads: list[FormDownloadResponse] = []
    if _answer_has_supported_form_context(answer):
        form_downloads = _selected_form_downloads(selected_form_names, available_forms)
    return QueryResponse(
        answer=answer,
        citations=citations,
        form_downloads=form_downloads,
        conversation_id=conversation_id,
    )


# FAQ pinned statis yang selalu tampil paling atas.
# Item ini tidak disimpan di faqs.json dan tidak bisa diedit/dihapus biasa.
# Hanya gambarnya yang bisa diganti lewat POST /api/admin/faq-image.
PINNED_IMAGE_STEM = "organogram"
PINNED_IMAGE_EXTENSIONS = {".webp", ".png", ".jpg", ".jpeg"}


def _pinned_image_name() -> str:
    # Kembalikan nama file gambar FAQ pinned yang aktif.
    for extension in (".webp", ".png", ".jpg", ".jpeg"):
        if (ASSETS_DIR / f"{PINNED_IMAGE_STEM}{extension}").exists():
            return f"{PINNED_IMAGE_STEM}{extension}"
    return f"{PINNED_IMAGE_STEM}.webp"


def _pinned_image_url() -> str:
    # Kembalikan URL cache-busted untuk gambar FAQ pinned.
    name = _pinned_image_name()
    path = ASSETS_DIR / name
    # Tambahkan cache-bust dari mtime agar browser reload setelah gambar diganti.
    version = int(path.stat().st_mtime) if path.exists() else 0
    return f"/assets/{name}?v={version}"


def _pinned_faq_items() -> list[FAQItem]:
    # Buat kartu FAQ statis untuk entri organogram.
    return [
        FAQItem(
            id="",
            question="Bagaimana struktur organisasi ICS Compute?",
            answer="Berikut struktur organisasi (organogram) ICS Compute.",
            suggested_query="Bagaimana struktur organisasi ICS Compute?",
            image_url=_pinned_image_url(),
        ),
    ]


@app.get("/api/faq", response_model=list[FAQItem])
def get_faq() -> list[FAQItem]:
    # Kembalikan FAQ pinned lalu FAQ tersimpan.
    with FAQ_LOCK:
        stored = _load_faqs()
    return [*_pinned_faq_items(), *stored]


@app.post("/api/admin/faq-image", response_model=AdminDocumentResponse)
def upload_pinned_faq_image(
    payload: AdminDocumentPayload,
    authorization: str = Header(default=""),
) -> AdminDocumentResponse:
    # Ganti aset gambar organogram pinned.
    _require_admin(authorization)
    extension = Path(unquote(payload.filename)).suffix.lower()
    if extension not in PINNED_IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Gambar harus berformat .webp, .png, atau .jpg.",
        )

    content = _decode_document(payload.content_base64)
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)

    # Simpan hanya satu gambar organogram: hapus varian ekstensi lain dulu.
    for other_extension in PINNED_IMAGE_EXTENSIONS:
        if other_extension == extension:
            continue
        stale = ASSETS_DIR / f"{PINNED_IMAGE_STEM}{other_extension}"
        if stale.exists():
            stale.unlink()

    (ASSETS_DIR / f"{PINNED_IMAGE_STEM}{extension}").write_bytes(content)
    return AdminDocumentResponse(message="Gambar FAQ diperbarui.")


@app.post("/api/admin/faq", response_model=AdminFAQResponse)
def create_faq(
    payload: AdminFAQPayload,
    authorization: str = Header(default=""),
) -> AdminFAQResponse:
    # Buat lalu simpan FAQ baru.
    _require_admin(authorization)
    item = _build_faq_item(payload)
    with FAQ_LOCK:
        items = _load_faqs()
        items.append(item)
        _save_faqs(items)
    return AdminFAQResponse(message="FAQ inserted.", item=item)


@app.put("/api/admin/faq/{faq_id}", response_model=AdminFAQResponse)
def update_faq(
    faq_id: str,
    payload: AdminFAQPayload,
    authorization: str = Header(default=""),
) -> AdminFAQResponse:
    # Buat ulang lalu ganti FAQ yang sudah ada.
    _require_admin(authorization)
    with FAQ_LOCK:
        items = _load_faqs()
        index = _find_faq_index(items, faq_id)
        item = _build_faq_item(payload, faq_id=items[index].id)
        items[index] = item
        _save_faqs(items)
    return AdminFAQResponse(message="FAQ updated.", item=item)


@app.delete("/api/admin/faq/{faq_id}", response_model=AdminFAQResponse)
def delete_faq(
    faq_id: str,
    authorization: str = Header(default=""),
) -> AdminFAQResponse:
    # Hapus satu FAQ tersimpan berdasarkan ID.
    _require_admin(authorization)
    with FAQ_LOCK:
        items = _load_faqs()
        index = _find_faq_index(items, faq_id)
        items.pop(index)
        _save_faqs(items)
    return AdminFAQResponse(message="FAQ deleted.")


@app.get("/api/library", response_model=list[LibraryItem])
def get_library(authorization: str = Header(default="")) -> list[LibraryItem]:
    # Kembalikan daftar library dokumen admin.
    _require_admin(authorization)
    return _iter_library_items()


@app.post("/api/admin/documents", response_model=AdminDocumentResponse)
def save_document(
    payload: AdminDocumentPayload,
    authorization: str = Header(default=""),
) -> AdminDocumentResponse:
    # Tambahkan atau ganti dokumen backend yang dikelola.
    _require_admin(authorization)
    data_dir = _get_data_dir().resolve()
    data_dir.mkdir(parents=True, exist_ok=True)

    filename = Path(unquote(payload.filename)).name
    suffix = Path(filename).suffix.lower()
    if suffix not in LIBRARY_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported document type.")

    content = _decode_document(payload.content_base64)

    if payload.replace_path:
        target_path = _resolve_document_path(payload.replace_path)
        if not target_path.exists() or not target_path.is_file():
            raise HTTPException(status_code=404, detail="Document not found.")
        if target_path.suffix.lower() != suffix:
            raise HTTPException(
                status_code=400,
                detail="Replacement file type must match the existing document.",
            )
        action = "updated"
    else:
        target_path = (data_dir / filename).resolve()
        try:
            target_path.relative_to(data_dir)
        except ValueError as error:
            raise HTTPException(status_code=400, detail="Invalid filename.") from error
        if target_path.exists():
            raise HTTPException(status_code=409, detail="Document already exists.")
        action = "inserted"

    target_path.write_bytes(content)
    requires_reindex = _is_embeddable_path(target_path)
    return AdminDocumentResponse(
        message=f"Document {action}.",
        requires_reindex=requires_reindex,
        item=_to_library_item(target_path, data_dir),
    )


@app.delete("/api/admin/documents/{document_path:path}", response_model=AdminDocumentResponse)
def delete_document(
    document_path: str,
    authorization: str = Header(default=""),
) -> AdminDocumentResponse:
    # Hapus satu dokumen terkelola dan laporkan kebutuhan reindex.
    _require_admin(authorization)
    target_path = _resolve_document_path(document_path)

    if not target_path.exists() or not target_path.is_file():
        raise HTTPException(status_code=404, detail="Document not found.")
    if target_path.suffix.lower() not in LIBRARY_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported document type.")

    requires_reindex = _is_embeddable_path(target_path)
    target_path.unlink()
    return AdminDocumentResponse(message="Document deleted.", requires_reindex=requires_reindex)


@app.post("/api/admin/reindex", response_model=AdminReindexResponse)
def reindex_documents(authorization: str = Header(default="")) -> AdminReindexResponse:
    # Bangun ulang vector database dari dokumen sumber saat ini.
    _require_admin(authorization)
    if not REINDEX_LOCK.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Reindex is already running.")

    try:
        from backend.preprocessing.ingest import main as rebuild_knowledge_base

        rebuild_knowledge_base()
    except Exception as error:
        raise HTTPException(
            status_code=500,
            detail=f"Rebuild embeddings failed: {error}",
        ) from error
    finally:
        REINDEX_LOCK.release()

    return AdminReindexResponse(message="Embeddings rebuilt.")


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# Cocokkan placeholder bracket yang dipakai template form, misalnya "[  ]".
FORM_PLACEHOLDER_PATTERN = re.compile(r"\[[^\[\]]*\]")
# Cell "field" adalah cell yang seluruh isinya hanya satu bracket.
# Placeholder inline seperti "Nama: [ ]" atau header dilewati agar form tetap singkat.
FORM_FIELD_CELL_PATTERN = re.compile(r"^\[[^\[\]]*\]$")


class FormFillPayload(BaseModel):
    path: str = Field(..., min_length=1)
    values: dict[str, str] = Field(default_factory=dict)


def _field_label(worksheet, cell) -> str | None:
    """Find a clean text label for a placeholder cell, or None if there isn't one.

    Looks inside the bracket, then at the nearest text to the left, then above.
    Returns None when no proper label exists (e.g. a stray grid cell), so it can
    be skipped instead of showing a "[  ]" field.
    """
    inside = str(cell.value).strip()[1:-1].strip()
    if inside:
        return inside
    for column in range(cell.column - 1, 0, -1):
        left = worksheet.cell(row=cell.row, column=column).value
        if isinstance(left, str):
            candidate = left.strip().rstrip(":").strip()
            if candidate and not FORM_PLACEHOLDER_PATTERN.search(candidate):
                return candidate
    if cell.row > 1:
        above = worksheet.cell(row=cell.row - 1, column=cell.column).value
        if isinstance(above, str):
            candidate = above.strip().rstrip(":").strip()
            if candidate and not FORM_PLACEHOLDER_PATTERN.search(candidate):
                return candidate
    return None


def _scan_form_fields(path: Path) -> list[dict[str, str]]:
    """List the main info fields of a form: the first contiguous block of
    labelled "[  ]" rows in each sheet. Later sections (free text, cost tables,
    signature blocks) are skipped so the fill form stays short."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    fields: list[dict[str, str]] = []
    for index, worksheet in enumerate(workbook.worksheets):
        started = False
        for row in worksheet.iter_rows():
            row_fields: list[dict[str, str]] = []
            for cell in row:
                if not (isinstance(cell.value, str) and FORM_FIELD_CELL_PATTERN.match(cell.value.strip())):
                    continue
                label = _field_label(worksheet, cell)
                if label:
                    row_fields.append(
                        {"key": f"{index}:{cell.coordinate}", "label": label}
                    )
            if row_fields:
                started = True
                fields.extend(row_fields)
            elif started:
                break  # end of the top info block for this sheet
    return fields


def _unique_form_fields(path: Path) -> list[dict[str, str]]:
    """Deduplicate the scanned fields by label so a repeated block (e.g. the
    same info section on 3 sheets) shows as one input."""
    seen: set[str] = set()
    unique: list[dict[str, str]] = []
    for field in _scan_form_fields(path):
        if field["label"] in seen:
            continue
        seen.add(field["label"])
        unique.append({"key": field["label"], "label": field["label"]})
    return unique


def _fill_form_placeholders(path: Path, values: dict[str, str]) -> bytes:
    """Fill every placeholder cell whose label the user provided a value for.

    Values are keyed by label, so one entry fills all cells sharing that label
    (e.g. the same field repeated across sheets). Blanks are left untouched.
    """
    from openpyxl import load_workbook

    coords_by_label: dict[str, list[str]] = {}
    for field in _scan_form_fields(path):
        coords_by_label.setdefault(field["label"], []).append(field["key"])

    workbook = load_workbook(path)
    for label, raw_value in values.items():
        value = str(raw_value).strip()
        if not value:
            continue
        for coord_key in coords_by_label.get(label, []):
            try:
                index_str, coordinate = coord_key.split(":", 1)
                cell = workbook.worksheets[int(index_str)][coordinate]
            except (ValueError, IndexError, KeyError):
                continue
            if isinstance(cell.value, str) and FORM_PLACEHOLDER_PATTERN.search(cell.value):
                cell.value = FORM_PLACEHOLDER_PATTERN.sub(lambda _match: value, cell.value, count=1)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _resolve_form_path(path: str) -> Path:
    # Tentukan dan validasi bahwa path menunjuk ke template form yang bisa diisi.
    resolved_path = _resolve_document_path(path)
    if not resolved_path.exists() or not resolved_path.is_file():
        raise HTTPException(status_code=404, detail="Form not found.")
    if (
        resolved_path.suffix.lower() != ".xlsx"
        or _document_kind_for_path(resolved_path) != "form"
    ):
        raise HTTPException(status_code=400, detail="Dokumen ini bukan form yang bisa diisi.")
    return resolved_path


@app.get("/api/documents/{document_path:path}")
def download_document(
    document_path: str,
    authorization: str = Header(default=""),
) -> FileResponse:
    # Unduh dokumen tersimpan, dengan file non-form dibatasi untuk admin.
    resolved_path = _resolve_document_path(document_path)

    if not resolved_path.exists() or not resolved_path.is_file():
        raise HTTPException(status_code=404, detail="Document not found.")
    if _document_kind_for_path(resolved_path) != "form":
        _require_admin(authorization)

    return FileResponse(
        path=resolved_path,
        filename=resolved_path.name,
        headers={"Cache-Control": "no-store"},
    )


@app.get("/api/forms/fields")
def form_fields(path: str) -> dict[str, object]:
    # Tampilkan field input sederhana yang terdeteksi di template form.
    resolved_path = _resolve_form_path(path)
    return {"fields": _unique_form_fields(resolved_path)}


@app.post("/api/forms/fill")
def fill_form(payload: FormFillPayload) -> Response:
    # Isi template form di memory lalu kembalikan file xlsx hasilnya.
    resolved_path = _resolve_form_path(payload.path)
    content = _fill_form_placeholders(resolved_path, payload.values)
    nama = next(
        (v.strip() for v in payload.values.values() if v.strip()),
        "",
    )
    suffix = f" - {nama}" if nama else ""
    filename = f"{resolved_path.stem}{suffix}.xlsx"
    return Response(
        content=content,
        media_type=XLSX_MIME,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


@app.get("/", response_class=FileResponse, include_in_schema=False)
def frontend_app() -> FileResponse:
    # Sajikan file index frontend untuk route root.
    index_file = FRONTEND_DIR / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=404, detail="Frontend bundle not found.")
    return FileResponse(index_file)
